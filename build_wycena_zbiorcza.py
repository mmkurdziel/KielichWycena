# -*- coding: utf-8 -*-
"""Generator profesjonalnego, zbiorczego pliku wyceny 'Kielich'."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# ----------------------------------------------------------------------------
# PALETA / STYLE
# ----------------------------------------------------------------------------
NAVY   = "1F3864"   # tytuly
BLUE   = "2E5496"   # naglowki kolumn
STEEL  = "8EAADB"   # naglowki sekcji
LIGHT  = "D6DCE5"   # wiersze pomocnicze / parametry
GOLD   = "FFE699"   # wiersze RAZEM / SUMA
GREEN  = "C6E0B4"   # wartosci rekomendowane / sumy modulow
GREY   = "F2F2F2"   # zebra

WHITE_BOLD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
HDR_FONT   = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BOLD       = Font(name="Calibri", size=10, bold=True)
NORM       = Font(name="Calibri", size=10)
SMALL      = Font(name="Calibri", size=9, italic=True, color="404040")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
LEFT_T = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hexc):
    return PatternFill("solid", fgColor=hexc)

PLN = '#,##0 "PLN"'
HRS = '#,##0 "h"'
PCT = '0%'

wb = openpyxl.Workbook()

# ============================================================================
# 1. LEGENDA
# ============================================================================
ws = wb.active
ws.title = "Legenda"
ws.sheet_view.showGridLines = False
widths = [3, 34, 20, 60]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

def title_bar(ws, text, last_col=4, row=1, fillc=NAVY):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font = TITLE_FONT; c.fill = fill(fillc); c.alignment = LEFT
    ws.row_dimensions[row].height = 26
    return row

def section(ws, text, r, last_col=4, fillc=STEEL):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    c = ws.cell(row=r, column=1, value=text)
    c.font = WHITE_BOLD; c.fill = fill(fillc); c.alignment = LEFT
    ws.row_dimensions[r].height = 20
    return r + 1

title_bar(ws, 'WYCENA APLIKACJI "KIELICH" — symulator monitora pacjenta (iPad · .NET MAUI)')
ws.merge_cells("A2:D2")
c = ws.cell(2, 1, "Zbiorcze zestawienie wariantów wyceny · aplikacja szkoleniowa (edukacyjna, NIE wyrób medyczny) · Pilot ↔ Monitor · offline P2P")
c.font = SMALL; c.alignment = LEFT
ws.merge_cells("A3:D3")
c = ws.cell(3, 1, "Data wyceny: 2026-06-15   ·   Wszystkie kwoty = Godziny × Stawka (parametry edytowalne poniżej)")
c.font = SMALL; c.alignment = LEFT

r = 5
r = section(ws, "PARAMETRY WYCENY (edytowalne — zmiana automatycznie przelicza wszystkie arkusze)", r)
params = [
    ("Stawka godzinowa (orientacyjna)", 50, "PLN/h", "Bazowa, orientacyjna stawka rozliczeniowa. Edytuj tę komórkę, aby przeliczyć całą wycenę."),
    ("Bufor ryzyka / rezerwa", 0.15, "%",  "15% od sumy godzin modułów (nieprzewidziane, iteracje, ryzyka techniczne)."),
    ("Kurs USD → PLN", 4.05, "USD/PLN", "Do przeliczenia kosztu Apple Developer Program (99 USD/rok)."),
]
hdrs = ["", "Parametr", "Wartość", "Opis"]
for j, h in enumerate(hdrs, 1):
    cc = ws.cell(r, j, h); cc.font = HDR_FONT; cc.fill = fill(BLUE); cc.alignment = CENTER; cc.border = BORDER
r += 1
param_rows = {}
for name, val, unit, desc in params:
    ws.cell(r, 2, name).font = BOLD
    vc = ws.cell(r, 3, val); vc.font = BOLD; vc.alignment = CENTER; vc.fill = fill(GOLD)
    if unit == "%":
        vc.number_format = "0%"
    ws.cell(r, 4, f"{desc}  [{unit}]").font = NORM
    for j in range(2, 5):
        ws.cell(r, j).border = BORDER
    ws.cell(r, 4).alignment = LEFT
    param_rows[name] = r
    r += 1

# Named ranges -> uzywane w formulach we wszystkich arkuszach
def add_name(nm, row):
    ref = f"Legenda!$C${row}"
    try:
        wb.defined_names.add(DefinedName(nm, attr_text=ref))
    except Exception:
        wb.defined_names[nm] = DefinedName(nm, attr_text=ref)

add_name("Stawka", param_rows["Stawka godzinowa (orientacyjna)"])
add_name("Bufor_proc", param_rows["Bufor ryzyka / rezerwa"])
add_name("KursUSD", param_rows["Kurs USD → PLN"])

r += 1
r = section(ws, "LEGENDA WARIANTÓW WYCENY", r)
var_hdr = ["", "Wariant (arkusz)", "Krótko", "Charakterystyka"]
for j, h in enumerate(var_hdr, 1):
    cc = ws.cell(r, j, h); cc.font = HDR_FONT; cc.fill = fill(BLUE); cc.alignment = CENTER; cc.border = BORDER
r += 1
variants = [
    ("V1 — Reuse (1:1)", "reuse, bez bufora",
     "Ponowne użycie biblioteki Infirmary Integrated. Najtańszy, najbardziej optymistyczny. Bez rezerwy ryzyka — łatwo przekroczyć budżet."),
    ("V2 — Własny kod", "od zera, bez bufora",
     "Wszystkie algorytmy pisane od zera (clean-room). Warianty A–D (prosty→pełny) oraz opcje sterowania 1:N."),
    ("V3 — Realistyczna (A/B/C)", "własny + bufor",
     "Pełny zakres ze szczegółowym rozbiciem na czynności. Bufor ryzyka 15%. A=podstawowy, B=+1:N, C=+reuse."),
    ("V4 — Zunifikowany ★", "reuse + bufor + sprzęt",
     "REKOMENDOWANA oferta: zakres V3 + oszczędność reuse + bufor ryzyka + koszty sprzętu. Pełna, realistyczna cena."),
]
for name, short, desc in variants:
    ws.cell(r, 2, name).font = BOLD
    ws.cell(r, 3, short).font = NORM; ws.cell(r, 3).alignment = LEFT
    ws.cell(r, 4, desc).font = NORM; ws.cell(r, 4).alignment = LEFT
    for j in range(2, 5):
        ws.cell(r, j).border = BORDER
    if name.endswith("★"):
        for j in range(2, 5):
            ws.cell(r, j).fill = fill(GREEN)
    r += 1

r += 1
r = section(ws, "SŁOWNIK MODUŁÓW (M1–M12) — kolejność realizacji", r)
mod_hdr = ["", "Moduł", "Nazwa", "Zakres"]
for j, h in enumerate(mod_hdr, 1):
    cc = ws.cell(r, j, h); cc.font = HDR_FONT; cc.fill = fill(BLUE); cc.alignment = CENTER; cc.border = BORDER
r += 1
modules_gloss = [
    ("M1", "Architektura i komunikacja P2P", "Warstwy/MVVM, abstrakcja transportu, PoC Multipeer Connectivity, synchronizacja stanu, reconnect."),
    ("M2", "Widok Monitora (przebiegi)", "Silnik renderowania przebiegów, EKG/SpO₂/BP/CVP/PAP/EtCO₂, kolumna parametrów liczbowych."),
    ("M3", "Pilot: pulpit instruktora", "Layout Pilota, ręczna zmiana parametrów, wybór rytmów/zakłóceń, integracja synchronizacji."),
    ("M4", "Arytmie i zakłócenia", "17 stanów EKG + EtCO₂/PAP/CVP/arterial, mechanizm przełączania, testy wiarygodności wizualnej."),
    ("M5", "System scenariuszy", "Model danych, zapis lokalny, CRUD, edytor, szybkie przełączanie w trakcie szkolenia."),
    ("M6", "Tryb Defib", "Defibrylacja: panel POWER/ENERGY/CHARGE/SHOCK, wybór energii, efekty, integracja z logiką rytmu."),
    ("M7", "Tryb AED", "Flow analiza→ładowanie→shock, komunikaty głosowe, integracja zmiany rytmu."),
    ("M8", "Kardiowersja zsynchronizowana", "Tryb SYNC, marker R-wave, ładowanie/wyładowanie zsynchronizowane."),
    ("M9", "Stymulacja (Pacing)", "Rate/output/mA, wizualizacja pobudzeń, model capture/no-capture (uproszczony; pełny = Opcja B)."),
    ("M10", "Logika wyładowań i zmiany rytmu", "Silnik reguł (po N wyładowaniach → rytm), konfiguracja w Pilocie, wspólne dla Defib/AED."),
    ("M11", "Ustawienia", "Aplikacja, audio, metronom (RKO), domyślne trybów terapii."),
    ("M12", "Testy, stabilizacja, dystrybucja", "Testy e2e na 2 iPadach, przypadki brzegowe, UX, TestFlight, dokumentacja."),
]
for m, nm, sc in modules_gloss:
    ws.cell(r, 2, m).font = BOLD; ws.cell(r, 2).alignment = CENTER
    ws.cell(r, 3, nm).font = BOLD; ws.cell(r, 3).alignment = LEFT
    ws.cell(r, 4, sc).font = NORM; ws.cell(r, 4).alignment = LEFT
    for j in range(2, 5):
        ws.cell(r, j).border = BORDER
    r += 1

r += 1
# ============================================================================
# SŁOWNIK POJĘĆ — medycznych, symulatora i technicznych
# ============================================================================
import math
r = section(ws, "SŁOWNIK POJĘĆ — medycznych, symulatora i technicznych", r)
cc = ws.cell(r, 2, "Pojęcie / zwrot"); cc.font = HDR_FONT; cc.fill = fill(BLUE); cc.alignment = CENTER; cc.border = BORDER
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
cc = ws.cell(r, 3, "Wyjaśnienie"); cc.font = HDR_FONT; cc.fill = fill(BLUE); cc.alignment = CENTER
ws.cell(r, 3).border = BORDER; ws.cell(r, 4).border = BORDER
r += 1

glossary = [
    ("CAT", "Warianty wyceny"),
    ("T", "Wariant „własny” (clean-room)",
     "Cały kod aplikacji — w tym silnik fal i fizjologia — pisany od zera, bez gotowej biblioteki. Większy nakład i ryzyko, ale pełna kontrola i brak zależności licencyjnych."),
    ("T", "Wariant „reuse”",
     "Ponowne użycie open-source’owej biblioteki Infirmary Integrated (Apache 2.0) dla najtrudniejszych części (rytmy, fale, fizjologia). Taniej i szybciej; wymaga atrybucji licencyjnej."),
    ("T", "„prosty” vs „pełny” (w wariancie własnym)",
     "Nie różnią się zakresem funkcji, tylko poziomem zaawansowania 3 modułów (przebiegi, fizjologia, pacing). „Prosty” = uproszczone kształty fal i brak korelacji parametrów (taniej, mniej wiarygodne). „Pełny” = matematyczne generatory fal i wzajemnie powiązana fizjologia (drożej, realistyczniej)."),

    ("CAT", "Stymulacja serca (pacing)"),
    ("T", "Pacing (stymulacja)",
     "Tryb urządzenia wysyłający rytmiczne impulsy elektryczne, by wymusić skurcze serca przy zbyt wolnym rytmie (bradykardia, blok). Instruktor ustawia częstość i siłę impulsu."),
    ("T", "rate",
     "Częstość stymulacji — ile impulsów na minutę wysyła stymulator."),
    ("T", "output / mA",
     "Siła (natężenie) impulsu w miliamperach — decyduje, czy serce odpowie na stymulację."),
    ("T", "Spike (impuls stymulacji)",
     "Cienka, pionowa kreska na zapisie EKG oznaczająca moment, w którym stymulator wysłał impuls elektryczny."),
    ("T", "Wizualizacja spikes",
     "Rysowanie tych pików na przebiegu EKG we właściwym tempie — kursant widzi na monitorze, że stymulator pracuje."),
    ("T", "Capture (wychwycenie) / no-capture",
     "Capture = po impulsie stymulatora serce faktycznie odpowiada skurczem (na EKG pojawia się pobudzenie). No-capture = jest spike, ale serce nie reaguje (np. za mały prąd) — stymulacja nieskuteczna."),
    ("T", "Przełącznik capture",
     "W wersji uproszczonej: ręczny przełącznik dla instruktora „pokaż wychwycenie / brak wychwycenia”. W wersji pełnej (Opcja B) aplikacja sama wylicza wychwycenie z progu (ustawione mA vs próg pobudzenia)."),

    ("CAT", "Terapie elektryczne"),
    ("T", "Defibrylacja",
     "Wyładowanie elektryczne (niezsynchronizowane) przerywające groźne arytmie (VF, VT bez tętna). Panel: POWER, ENERGY, CHARGE, SHOCK, DISARM; energia 10–360 J."),
    ("T", "Kardiowersja zsynchronizowana (tryb SYNC)",
     "Wyładowanie zsynchronizowane z załamkiem R serca, stosowane w arytmiach z tętnem, by nie trafić w „fazę wrażliwą” cyklu serca."),
    ("T", "Marker R-wave",
     "Znacznik na EKG wskazujący załamek R (szczyt zespołu QRS) — punkt, z którym synchronizuje się wyładowanie w trybie SYNC."),
    ("T", "AED",
     "Automatyczny defibrylator zewnętrzny — prowadzi ratownika krok po kroku: analiza rytmu → „shock advised / no shock” → ładowanie → wyładowanie, z komunikatami głosowymi."),
    ("T", "Logika wyładowań",
     "Konfigurowalne reguły symulatora: „po N wyładowaniach zmień rytm na X” — pozwala instruktorowi budować scenariusze reakcji pacjenta."),

    ("CAT", "Sygnały i parametry monitora"),
    ("T", "EKG (ECG)",
     "Zapis czynności elektrycznej serca — podstawowy przebieg na monitorze."),
    ("T", "Arytmie / rytmy (17 stanów)",
     "Nieprawidłowe rytmy serca odwzorowane w aplikacji: m.in. bloki, AF (migotanie przedsionków), flutter, junctional, LBBB/RBBB (bloki odnóg pęczka Hisa), ST elevation, SVT, Torsades, VF (migotanie komór), VT (częstoskurcz komorowy), Noise (zakłócenia)."),
    ("T", "HR",
     "Częstość akcji serca (uderzenia na minutę)."),
    ("T", "SpO₂ / pleth",
     "SpO₂ = wysycenie krwi tlenem (%); pleth = krzywa pletyzmograficzna z pulsoksymetru."),
    ("T", "BP (ciśnienie tętnicze)",
     "Krzywa i wartości ciśnienia tętniczego krwi."),
    ("T", "CVP",
     "Ośrodkowe ciśnienie żylne."),
    ("T", "PAP",
     "Ciśnienie w tętnicy płucnej."),
    ("T", "EtCO₂ (kapnografia)",
     "Stężenie dwutlenku węgla w wydychanym powietrzu — krzywa oddechowa."),
    ("T", "RR",
     "Częstość oddechów (oddechy na minutę)."),
    ("T", "Metronom (RKO)",
     "Dźwiękowy/wizualny wskaźnik tempa uciśnięć klatki piersiowej podczas resuscytacji krążeniowo-oddechowej."),

    ("CAT", "Pojęcia techniczne"),
    ("T", "Pilot ↔ Monitor",
     "Dwie role aplikacji: Pilot (urządzenie instruktora) steruje; Monitor (urządzenie kursanta) wyświetla obraz monitora pacjenta."),
    ("T", "P2P (peer-to-peer)",
     "Bezpośrednie połączenie urządzeń bez internetu/serwera — działa offline w sali szkoleniowej."),
    ("T", "Multipeer Connectivity",
     "Technologia Apple do łączności P2P po Wi-Fi/Bluetooth między urządzeniami iOS."),
    ("T", "1:1 vs 1:N",
     "1:1 = jeden Pilot steruje jednym Monitorem. 1:N = jeden Pilot steruje wieloma Monitorami (do 4)."),
    ("T", "broadcast",
     "Wysyłka tych samych parametrów jednocześnie do wszystkich podłączonych Monitorów (tryb 1:N)."),
    ("T", ".NET MAUI",
     "Wieloplatformowy framework C# (iOS/Android/desktop z jednego kodu) — technologia, w której pisana jest aplikacja."),
    ("T", "SkiaSharp",
     "Biblioteka grafiki 2D używana do rysowania przebiegów/fal na ekranie."),
    ("T", "TestFlight",
     "Usługa Apple do dystrybucji aplikacji testowej na wybrane urządzenia, bez publikacji w App Store."),
    ("T", "Infirmary Integrated (reuse)",
     "Otwarta (Apache 2.0) biblioteka symulatora pacjenta, z której można ponownie użyć rytmów, fal i fizjologii."),
]
for item in glossary:
    if item[0] == "CAT":
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c = ws.cell(r, 2, item[1]); c.font = WHITE_BOLD; c.fill = fill(STEEL); c.alignment = LEFT
        ws.row_dimensions[r].height = 18
    else:
        _, term, expl = item
        tc = ws.cell(r, 2, term); tc.font = BOLD; tc.alignment = LEFT_T; tc.border = BORDER
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ec = ws.cell(r, 3, expl); ec.font = NORM; ec.alignment = LEFT_T
        ws.cell(r, 3).border = BORDER; ws.cell(r, 4).border = BORDER
        ws.row_dimensions[r].height = max(16, math.ceil(len(expl) / 85.0) * 14)
    r += 1

r += 1
r = section(ws, "OZNACZENIA KOLORÓW", r)
legend_colors = [
    (GOLD,  "Wiersze RAZEM / SUMA / parametry edytowalne"),
    (GREEN, "Wartości rekomendowane / sumy modułów"),
    (LIGHT, "Wiersze pomocnicze, korekty, opcje"),
    (BLUE,  "Nagłówki kolumn (biały tekst)"),
    (STEEL, "Nagłówki sekcji (biały tekst)"),
]
for col, desc in legend_colors:
    cc = ws.cell(r, 2, ""); cc.fill = fill(col); cc.border = BORDER
    ws.cell(r, 3, "").border = BORDER
    ws.cell(r, 4, desc).font = NORM; ws.cell(r, 4).alignment = LEFT
    r += 1

r += 1
r = section(ws, "ZAŁOŻENIA I UWAGI", r)
notes = [
    "Technologia: .NET MAUI (C#), Visual Studio / VS Code, SkiaSharp, Multipeer Connectivity, SQLite. Architektura przygotowana pod przyszły Android.",
    "Reuse: Infirmary Integrated (Apache 2.0) — wymagana atrybucja i plik NOTICE; UI i komunikacja P2P pisane od zera.",
    "Mac jako host budowania iOS jest WYMAGANY (kompilacja i podpisywanie). Założono, że klient posiada Maca (0 PLN).",
    "Fizyczne iPady wymagane TYLKO do testów komunikacji P2P (Wi-Fi/Multipeer). Emulator Xcode wystarcza do reszty developmentu.",
    "Dystrybucja: TestFlight / wewnętrzna. Apple Developer Program 99 USD/rok.",
    "Bufor ryzyka liczony od godzin modułów danego wariantu. Koszty stałe (sprzęt/konta) identyczne we wszystkich wariantach.",
    "Programista: 1 osoba, C#/.NET, bez doświadczenia z Apple/iOS — stąd istotny bufor ryzyka w wariantach V3/V4.",
    "Wycena NIE obejmuje instalacji środowiska/narzędzi — wyłącznie czas developmentu. Godziny szacunkowe.",
    "Opcje 1:N, pełny Pacing i port Android są poza sumami głównymi — patrz arkusz 'Opcje dodatkowe'.",
]
for n in notes:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    cc = ws.cell(r, 2, "• " + n); cc.font = NORM; cc.alignment = LEFT_T
    ws.row_dimensions[r].height = 28
    r += 1

# ============================================================================
# Pomocnicze do arkuszy wycen
# ============================================================================
def setup_sheet(name, widths):
    s = wb.create_sheet(name)
    s.sheet_view.showGridLines = False
    for i, w in enumerate(widths, 1):
        s.column_dimensions[get_column_letter(i)].width = w
    return s

def col_header(s, r, headers, fillc=BLUE):
    for j, h in enumerate(headers, 1):
        c = s.cell(r, j, h); c.font = HDR_FONT; c.fill = fill(fillc)
        c.alignment = CENTER; c.border = BORDER
    s.row_dimensions[r].height = 30
    return r + 1

# ============================================================================
# Arkusz: PODSUMOWANIE (zbiorcze porównanie)
# ============================================================================
ws = setup_sheet("Podsumowanie", [34, 13, 13, 13, 13, 13, 13, 13])
title_bar(ws, "PODSUMOWANIE — porównanie wszystkich wariantów wyceny", last_col=8)
ws.merge_cells("A2:H2")
ws.cell(2, 1, "Stawka i bufor ryzyka pobierane z arkusza 'Legenda'. Wartości godzinowe to dane wejściowe; kwoty liczone formułami.").font = SMALL

r = 4
headers = ["Pozycja", "V1\nreuse", "V2-A\nwłasny prosty", "V2-D\nwłasny pełny",
           "V3-A\nwłasny+buf", "V3-B\n+1:N", "V3-C\nreuse+buf", "V4 ★\nzunifikowany"]
r = col_header(ws, r, headers)

# wiersze: (etykieta, [V1,V2A,V2D,V3A,V3B,V3C,V4], typ)
hdev = ["Godziny modułów (development)", 312, 402, 488, 936, 1066, 824, 824]
data_rows = [
    ("Godziny modułów (development)", [312, 402, 488, 936, 1066, 824, 824], "hrs"),
    ("Bufor ryzyka (+15%)",           [0, 0, 0, 140, 160, 124, 124], "hrs"),
    ("RAZEM godziny",                 [312, 402, 488, 1076, 1226, 948, 948], "hrs_sum"),
    ("Koszt robocizny (PLN)",         [15600, 20100, 24400, 53800, 61300, 47400, 47400], "pln"),
    ("Koszty stałe (sprzęt/Apple)",   [7300, 0, 0, 5701, 5701, 5701, 5701], "pln"),
    ("SUMA CAŁKOWITA (PLN)",          [22900, 20100, 24400, 59501, 67001, 53101, 53101], "pln_sum"),
]
for label, vals, typ in data_rows:
    ws.cell(r, 1, label).font = BOLD if "RAZEM" in label or "SUMA" in label else NORM
    ws.cell(r, 1).alignment = LEFT; ws.cell(r, 1).border = BORDER
    for j, v in enumerate(vals, 2):
        c = ws.cell(r, j, v); c.alignment = CENTER; c.border = BORDER
        if typ in ("hrs", "hrs_sum"):
            c.number_format = HRS
        else:
            c.number_format = PLN
        if typ in ("hrs_sum",):
            c.fill = fill(LIGHT); c.font = BOLD
        if typ == "pln_sum":
            c.fill = fill(GOLD); c.font = BOLD
        if "RAZEM" in label or "SUMA" in label:
            c.font = BOLD
    # podswietl kolumne V4
    cc = ws.cell(r, 8); cc.fill = fill(GREEN if typ != "pln_sum" else GOLD)
    r += 1

# cechy
r += 1
r = col_header(ws, r, ["Cecha wariantu", "V1", "V2-A", "V2-D", "V3-A", "V3-B", "V3-C", "V4 ★"], fillc=STEEL)
feat = [
    ("Reuse Infirmary Integrated", ["TAK","nie","nie","nie","nie","TAK","TAK"]),
    ("Bufor ryzyka (15%)",         ["nie","nie","nie","TAK","TAK","TAK","TAK"]),
    ("Pełny zakres szczegółowy",   ["nie","nie","TAK","TAK","TAK","TAK","TAK"]),
    ("Połączenie 1:N (do 4)",      ["+16h","+20h","+35h","opcja","wliczone","opcja A","opcja A"]),
    ("Koszty sprzętu w SUMIE",     ["TAK","nie","nie","TAK","TAK","TAK","TAK"]),
    ("Port Android (orient.)",     ["+2000","brak","brak","~6000","~6000","~5500","~5500"]),
]
for label, vals in feat:
    ws.cell(r, 1, label).font = NORM; ws.cell(r, 1).alignment = LEFT; ws.cell(r, 1).border = BORDER
    for j, v in enumerate(vals, 2):
        c = ws.cell(r, j, v); c.alignment = CENTER; c.border = BORDER; c.font = NORM
    ws.cell(r, 8).fill = fill(GREEN)
    r += 1

r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
ws.cell(r, 1, "★ REKOMENDACJA: Wariant V4 — Zunifikowany (948 h · 53 101 PLN). Łączy szczegółowy zakres, oszczędność reuse, bufor ryzyka i koszty sprzętu — pełna, realistyczna oferta.").font = BOLD
ws.cell(r, 1).fill = fill(GREEN); ws.cell(r, 1).alignment = LEFT
ws.row_dimensions[r].height = 30
ws.freeze_panes = "B5"

# ============================================================================
# Generyczny budowniczy arkusza modulowego
# ============================================================================
def module_sheet(name, title, subtitle, headers, rows, totals_note=None,
                 hours_col=3, note_col=None):
    """rows: list of dict:
       {'type':'section','text':..}
       {'type':'mod','cells':[...], 'hours':int, 'note':..}
       {'type':'sum','label':.., 'from':r1,'to':r2}
    Zwraca arkusz. Kolumna kosztu = ostatnia przed uwagami liczona formula."""
    ncol = len(headers)
    widths = []
    return None

# Bardziej bezposrednie podejscie per-arkusz ponizej.

def build_simple_modules(sheetname, title, subtitle, modrows, rate_note=True,
                         extra_after=None, kicker=None):
    """modrows: list of tuples (nr, nazwa, zakres, godziny, uwaga)
       Tworzy tabele z kolumnami:
       Nr | Moduł | Zakres / zadania | Godziny | Koszt (PLN) | Uwagi
    """
    s = setup_sheet(sheetname, [7, 26, 48, 11, 14, 30])
    title_bar(s, title, last_col=6)
    s.merge_cells("A2:F2")
    s.cell(2, 1, subtitle).font = SMALL; s.cell(2, 1).alignment = LEFT
    r = 4
    r = col_header(s, r, ["Nr", "Moduł", "Zakres / zadania", "Godziny", "Koszt (PLN)", "Uwagi"])
    first = r
    zebra = False
    for nr, nm, zk, godz, uw in modrows:
        s.cell(r, 1, nr).font = BOLD; s.cell(r, 1).alignment = CENTER
        s.cell(r, 2, nm).font = BOLD; s.cell(r, 2).alignment = LEFT_T
        s.cell(r, 3, zk).font = NORM; s.cell(r, 3).alignment = LEFT_T
        gc = s.cell(r, 4, godz); gc.alignment = CENTER; gc.number_format = HRS
        kc = s.cell(r, 5, f"=D{r}*Stawka"); kc.alignment = CENTER; kc.number_format = PLN
        s.cell(r, 6, uw).font = SMALL; s.cell(r, 6).alignment = LEFT_T
        for j in range(1, 7):
            s.cell(r, j).border = BORDER
            if zebra:
                s.cell(r, j).fill = fill(GREY)
        s.row_dimensions[r].height = 30
        zebra = not zebra
        r += 1
    last = r - 1
    # SUMA
    s.cell(r, 2, "SUMA ROBOCIZNA (moduły)").font = BOLD
    s.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    s.cell(r, 2).alignment = LEFT
    gc = s.cell(r, 4, f"=SUM(D{first}:D{last})"); gc.number_format = HRS; gc.font = BOLD; gc.alignment = CENTER
    kc = s.cell(r, 5, f"=SUM(E{first}:E{last})"); kc.number_format = PLN; kc.font = BOLD; kc.alignment = CENTER
    for j in range(1, 7):
        s.cell(r, j).fill = fill(GREEN); s.cell(r, j).border = BORDER
    sum_row = r
    r += 2
    if kicker:
        for line in kicker:
            s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            s.cell(r, 1, line).font = NORM; s.cell(r, 1).alignment = LEFT_T
            s.row_dimensions[r].height = 26
            r += 1
    s.freeze_panes = "A5"
    return s, sum_row, r

# ============================================================================
# V1 — Reuse (1:1)
# ============================================================================
v1_mods = [
    ("M0", "Setup", "Xcode, MAUI iOS, certyfikaty, architektura, nauka MAUI", 30, ""),
    ("M1", "P2P 1:1", "Multipeer Connectivity binding, auto-discovery, sync, reconnect", 32, "Ryzyko: binding MAUI↔Apple API"),
    ("M2", "Przebiegi", "Port waveformów z Infirmary Integrated (SkiaSharp/MAUI)", 33, "Reuse open-source C#"),
    ("M3", "Fizjologia", "Port Physiology.cs + Rhythms.cs, 17 rytmów, stany zakłóceń", 28, "Reuse open-source C#"),
    ("M4", "Monitor UI", "Ciemny ekran, przebiegi + kolumna numeryczna, przełączanie widoków", 23, ""),
    ("M5", "Pilot UI", "Panel sterowania, parametry, rytmy, scenariusze, status", 27, ""),
    ("M6", "Scenariusze", "CRUD, SQLite, edytor, przełączanie w trakcie szkolenia", 22, "Adaptacja Scenario.cs"),
    ("M7", "Defib", "Defibrylacja + kardiowersja zsynchronizowana + pacing", 35, ""),
    ("M8", "AED", "Flow Analyze→Charge→Shock, efekty audio/wizualne", 13, ""),
    ("M9", "Logika wyładowań", "Zmiana rytmu po N wyładowaniach, łańcuchowanie reguł", 15, ""),
    ("M10", "Audio", "Port Audio.cs, beep HR, alarm, dźwięk wyładowania, metronom", 9, ""),
    ("M11", "Ustawienia", "Ogólne, audio, metronom, domyślne Defib/AED/pacing", 10, ""),
    ("M12", "Testy", "Emulator + fizyczne urządzenia P2P, optymalizacja, polish", 35, "⚠ Wymaga 2 fizycznych iPadów"),
]
build_simple_modules(
    "V1 — Reuse (1-1)",
    "WYCENA V1 — Reuse Infirmary Integrated (1 Pilot : 1 Monitor)",
    "Ponowne użycie biblioteki open-source · bez bufora ryzyka · najbardziej optymistyczny wariant",
    v1_mods,
    kicker=[
        "Oszczędność dzięki reuse II Library: ~108 h (~5 400 PLN). Moduły z reuse: Physiology.cs, Rhythms.cs, Waveform.*.cs, Scenario.cs, Audio.cs.",
        "Wariant 1:N (do wielu monitorów): +16 h w module M1 (P2P 1:wielu) → razem 328 h.",
        "Koszty stałe (poza robocizną): Mac mini 2 500 + 2× iPad 4 400 + Apple Dev 400 = 7 300 PLN. Port Android: +~2 000 PLN.",
    ],
)

# ============================================================================
# V2 — Własny kod (moduly stale + zmienne + 1:N + warianty)
# ============================================================================
s = setup_sheet("V2 — Własny kod", [7, 30, 12, 12, 13, 13, 26])
title_bar(s, "WYCENA V2 — Własny kod (clean-room, bez reuse)", last_col=7)
s.merge_cells("A2:G2")
s.cell(2, 1, "Wszystkie algorytmy pisane od zera · bez bufora ryzyka · warianty A–D (prosty→pełny) · sterowanie 1:N (Opcje A/B/C)").font = SMALL

r = 4
r = section(s, "MODUŁY STAŁE — niezależne od wariantu", r, last_col=7)
r = col_header(s, r, ["Nr", "Moduł", "Godz. 1:1", "Godz. 1:wielu", "Koszt 1:1 (PLN)", "Koszt 1:wielu (PLN)", "Uwagi"])
v2_fixed = [
    ("M0", "Setup (Xcode, MAUI, certyfikaty, architektura)", 32, 32, ""),
    ("M1", "P2P (Multipeer Connectivity binding)", 39, 59, "+20h dla 1:wielu (multi-peer broadcast)"),
    ("M4", "Monitor UI", 25, 25, ""),
    ("M5", "Pilot UI", 30, 30, ""),
    ("M6", "Scenariusze (SQLite, CRUD, edytor)", 27, 27, ""),
    ("M7", "Defib baza (defibrylacja + kardiowersja)", 29, 29, ""),
    ("M8", "AED", 14, 14, ""),
    ("M9", "Logika wyładowań", 17, 17, ""),
    ("M10", "Audio (od zera)", 18, 18, ""),
    ("M11", "Ustawienia", 11, 11, ""),
    ("M12", "Testy", 42, 42, ""),
]
first = r
zebra = False
for nr, nm, g1, gn, uw in v2_fixed:
    s.cell(r, 1, nr).font = BOLD; s.cell(r, 1).alignment = CENTER
    s.cell(r, 2, nm).font = NORM; s.cell(r, 2).alignment = LEFT_T
    s.cell(r, 3, g1).alignment = CENTER; s.cell(r, 3).number_format = HRS
    s.cell(r, 4, gn).alignment = CENTER; s.cell(r, 4).number_format = HRS
    s.cell(r, 5, f"=C{r}*Stawka").alignment = CENTER; s.cell(r, 5).number_format = PLN
    s.cell(r, 6, f"=D{r}*Stawka").alignment = CENTER; s.cell(r, 6).number_format = PLN
    s.cell(r, 7, uw).font = SMALL; s.cell(r, 7).alignment = LEFT_T
    for j in range(1, 8):
        s.cell(r, j).border = BORDER
        if zebra: s.cell(r, j).fill = fill(GREY)
    zebra = not zebra
    r += 1
last = r - 1
s.cell(r, 2, "SUMA moduły stałe").font = BOLD
for col, formula in [(3, f"=SUM(C{first}:C{last})"), (4, f"=SUM(D{first}:D{last})"),
                     (5, f"=SUM(E{first}:E{last})"), (6, f"=SUM(F{first}:F{last})")]:
    c = s.cell(r, col, formula); c.font = BOLD; c.alignment = CENTER
    c.number_format = HRS if col in (3, 4) else PLN
for j in range(1, 8):
    s.cell(r, j).fill = fill(GREEN); s.cell(r, j).border = BORDER
r += 2

r = section(s, "MODUŁY ZMIENNE — zależne od wariantu (wybierz wersję)", r, last_col=7)
r = col_header(s, r, ["Nr", "Moduł", "Wersja", "Godziny", "Koszt (PLN)", "—", "Uwagi"])
v2_var = [
    ("M2", "Przebiegi", "Prosta (lookup table, uproszczone kształty)", 48),
    ("M2", "Przebiegi", "Złożona (matematyczne generatory, wiarygodne)", 83),
    ("M3", "Fizjologia + arytmie", "Prosta (17 rytmów, bez korelacji)", 57),
    ("M3", "Fizjologia + arytmie", "Złożona (+ silnik korelacji, event-driven)", 97),
    ("M7", "Pacing", "Uproszczony (UI + spikes + przełącznik capture)", 13),
    ("M7", "Pacing", "Pełny (logika capture, fizjologiczna odpowiedź)", 24),
]
zebra = False
for nr, nm, wer, godz in v2_var:
    s.cell(r, 1, nr).font = BOLD; s.cell(r, 1).alignment = CENTER
    s.cell(r, 2, nm).font = NORM; s.cell(r, 2).alignment = LEFT_T
    s.cell(r, 3, wer).font = NORM; s.cell(r, 3).alignment = LEFT_T
    s.cell(r, 4, godz).alignment = CENTER; s.cell(r, 4).number_format = HRS
    s.cell(r, 5, f"=D{r}*Stawka").alignment = CENTER; s.cell(r, 5).number_format = PLN
    s.cell(r, 7, "").font = NORM
    for j in range(1, 8):
        s.cell(r, j).border = BORDER
        if zebra: s.cell(r, j).fill = fill(GREY)
    zebra = not zebra
    r += 1
r += 1

r = section(s, "EKRAN ZARZĄDZANIA URZĄDZENIAMI — tylko wersja 1:wielu", r, last_col=7)
r = col_header(s, r, ["Opcja", "Opis", "—", "Godziny", "Koszt (PLN)", "—", "Uwagi"])
v2_ctrl = [
    ("Opcja A", "broadcast — lista urządzeń + te same parametry do wszystkich", 12),
    ("Opcja B", "indywidualne — niezależne sterowanie każdym Monitorem", 35),
    ("Opcja C", "broadcast + override — domyślnie broadcast, możliwość nadpisania", 40),
]
for op, opis, godz in v2_ctrl:
    s.cell(r, 1, op).font = BOLD; s.cell(r, 1).alignment = CENTER
    s.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    s.cell(r, 2, opis).font = NORM; s.cell(r, 2).alignment = LEFT_T
    s.cell(r, 4, godz).alignment = CENTER; s.cell(r, 4).number_format = HRS
    s.cell(r, 5, f"=D{r}*Stawka").alignment = CENTER; s.cell(r, 5).number_format = PLN
    for j in range(1, 8):
        s.cell(r, j).border = BORDER
    r += 1
r += 1

r = section(s, "PODSUMOWANIE WARIANTÓW (robocizna, bez sprzętu)", r, last_col=7)
r = col_header(s, r, ["Wariant", "Sterowanie", "Godziny", "Koszt robocizny (PLN)", "+ Apple Dev (PLN)", "RAZEM (PLN)", "Uwagi"])
v2_sum = [
    ("1:1 A — prosty", "—", 402, 400),
    ("1:1 B", "—", 437, 400),
    ("1:1 C", "—", 453, 400),
    ("1:1 D — pełny", "—", 488, 400),
    ("1:N A — prosty", "Opcja A — broadcast", 434, 400),
    ("1:N A — prosty", "Opcja B — indywidualne", 457, 400),
    ("1:N A — prosty", "Opcja C — broadcast+override", 462, 400),
    ("1:N D — pełny", "Opcja A — broadcast", 520, 400),
    ("1:N D — pełny", "Opcja B — indywidualne", 543, 400),
    ("1:N D — pełny", "Opcja C — broadcast+override", 548, 400),
]
zebra = False
for war, ster, godz, apple in v2_sum:
    s.cell(r, 1, war).font = BOLD; s.cell(r, 1).alignment = LEFT
    s.cell(r, 2, ster).font = NORM; s.cell(r, 2).alignment = LEFT
    s.cell(r, 3, godz).alignment = CENTER; s.cell(r, 3).number_format = HRS
    s.cell(r, 4, f"=C{r}*Stawka").alignment = CENTER; s.cell(r, 4).number_format = PLN
    s.cell(r, 5, apple).alignment = CENTER; s.cell(r, 5).number_format = PLN
    s.cell(r, 6, f"=D{r}+E{r}").alignment = CENTER; s.cell(r, 6).number_format = PLN; s.cell(r, 6).font = BOLD
    s.cell(r, 6).fill = fill(GOLD)
    for j in range(1, 8):
        s.cell(r, j).border = BORDER
        if zebra and j != 6: s.cell(r, j).fill = fill(GREY)
    zebra = not zebra
    r += 1
r += 1
s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
s.cell(r, 1, "Port na Android (delta): +40 h (~2 000 PLN) — Wi-Fi Direct/NSD zamiast Multipeer + testy + poprawki UI.").font = NORM
s.cell(r, 1).alignment = LEFT
s.freeze_panes = "A5"

# ============================================================================
# V3 — Realistyczna (szczegolowe rozbicie na czynnosci)
# ============================================================================
s = setup_sheet("V3 — Realistyczna", [7, 50, 11, 14, 30])
title_bar(s, "WYCENA V3 — Realistyczna: szczegółowe rozbicie na czynności (Wariant A)", last_col=5)
s.merge_cells("A2:E2")
s.cell(2, 1, "Własny kod · pełny zakres · bufor ryzyka 15% · kolejność realizacji M1→M12 · kolumna korekt reuse dla Wariantu C").font = SMALL

r = 4
r = col_header(s, r, ["Nr", "Kamień milowy / czynność", "Godziny", "Koszt (PLN)", "Uwagi"])

# struktura: (milestone_name, [(czynnosc, godz, uwaga), ...])
v3 = [
    ("M1 — Architektura i komunikacja P2P", [
        ("Projekt architektury: warstwy, MVVM, abstrakcja transportu pod przyszły Android", 16, ""),
        ("Ekran startowy + wybór roli (Pilot / Monitor)", 8, ""),
        ("PoC Multipeer Connectivity w MAUI (binding warstwy natywnej, wykrywanie)", 24, "Ryzyko R1"),
        ("Warstwa synchronizacji stanu w czasie rzeczywistym (model, protokół, serializacja)", 24, ""),
        ("Lista dostępnych Pilotów + wybór bez wpisywania kodu", 10, ""),
        ("Obsługa rozłączeń / reconnect + wskaźnik stanu połączenia", 12, ""),
        ("Testy komunikacji na 2 fizycznych iPadach", 10, "Wymaga 2 iPadów"),
    ]),
    ("M2 — Widok Monitora (silnik przebiegów + parametry)", [
        ("Silnik renderowania przebiegów (rysowanie, animacja, skala czasu)", 30, "Główny driver kosztów"),
        ("Przebiegi bazowe: EKG, pleth/SpO₂, ciśnienie tętnicze, CVP/PAP, EtCO₂", 40, ""),
        ("Kolumna parametrów liczbowych (HR, SpO₂, BP, CVP, PAP, EtCO₂, RR, Temp, Timer)", 20, ""),
        ("Layout monitora (ciemne tło, czytelność z odległości, układ iPad)", 16, ""),
        ("Synchronizacja przebiegów i parametrów z Pilotem", 14, ""),
    ]),
    ("M3 — Pilot: pulpit instruktora", [
        ("Layout Pilota (podgląd monitora + panel sterowania)", 16, ""),
        ("Ręczna zmiana parametrów (kontrolki, walidacja, wysyłka do Monitora)", 24, ""),
        ("UI wyboru rytmów i zakłóceń", 16, ""),
        ("Podgląd stanu połączenia i statusów", 8, ""),
        ("Integracja z warstwą synchronizacji", 12, ""),
    ]),
    ("M4 — Arytmie i zakłócenia (stany przebiegów)", [
        ("EKG: 17 stanów (bloki, AF, flutter, junctional, LBBB/RBBB, ST, SVT, Torsades, VF, VT, Noise)", 60, "Wiarygodność = główny driver"),
        ("EtCO₂: 8 stanów", 16, ""),
        ("PAP / CVP: 7 stanów", 16, ""),
        ("Arterial / BP: 4 stany", 10, ""),
        ("Mechanizm przełączania stanów + wpływ na parametry + synchronizacja", 16, ""),
        ("Testy wiarygodności wizualnej", 12, ""),
    ]),
    ("M5 — System scenariuszy", [
        ("Model danych scenariusza (parametry, rytm, zakłócenia, terapie, notatki)", 12, ""),
        ("Zapis lokalny (persistencja danych na urządzeniu)", 12, ""),
        ("CRUD: tworzenie / edycja / kopiowanie / usuwanie", 24, ""),
        ("Edytor scenariuszy (interfejs)", 28, ""),
        ("Szybkie uruchamianie + przełączanie scenariuszy w trakcie szkolenia", 16, ""),
        ("Testy", 10, ""),
    ]),
    ("M6 — Tryb Defib (defibrylacja)", [
        ("Widok Defib + lokalne przełączanie widoków Monitor / Defib / AED", 16, ""),
        ("Panel wspólny (POWER, ENERGY, CHARGE, SHOCK, DISARM)", 16, ""),
        ("Wybór energii 10–360 J + animacja ładowania", 14, ""),
        ("SHOCK: efekt wizualny + dźwiękowy", 12, ""),
        ("Integracja z logiką zmiany rytmu", 8, ""),
        ("Testy", 8, ""),
    ]),
    ("M7 — Tryb AED", [
        ("Widok AED + logika prowadzenia (analiza, komunikaty, shock advised / no shock)", 28, ""),
        ("Integracja z logiką zmiany rytmu po wyładowaniach", 10, ""),
        ("Audio / komunikaty głosowe AED", 8, ""),
        ("Testy", 8, ""),
    ]),
    ("M8 — Kardiowersja zsynchronizowana", [
        ("Tryb SYNC + znacznik synchronizacji / marker R-wave", 16, ""),
        ("Ładowanie + wyładowanie zsynchronizowane", 12, ""),
        ("Efekty + integracja z logiką rytmu", 10, ""),
        ("Testy", 6, ""),
    ]),
    ("M9 — Stymulacja (Pacing) — model uproszczony", [
        ("Tryb pacing: rate, output / mA, start / stop", 18, ""),
        ("Wizualizacja pobudzeń stymulowanych na EKG", 16, ""),
        ("Uproszczony model capture / no-capture", 12, "Pełny model = Opcja B"),
        ("Testy", 6, ""),
    ]),
    ("M10 — Logika wyładowań i zmiany rytmu (konfigurowalna)", [
        ("Silnik reguł (po ilu wyładowaniach → jaki rytm; łańcuch reguł)", 20, ""),
        ("Konfiguracja w Pilocie + przypisanie do scenariusza", 16, ""),
        ("Wspólne dla Defib i AED", 8, ""),
        ("Testy", 8, ""),
    ]),
    ("M11 — Ustawienia (audio, metronom, terapie)", [
        ("Ustawienia aplikacji", 10, ""),
        ("Ustawienia audio", 8, ""),
        ("Metronom (RKO)", 12, ""),
        ("Ustawienia trybów terapii (Defib / AED / kardiowersja / pacing)", 10, ""),
    ]),
    ("M12 — Testy, stabilizacja, UX, dystrybucja", [
        ("Testy integracyjne end-to-end na 2 urządzeniach", 24, ""),
        ("Stabilizacja połączenia + przypadki brzegowe", 20, ""),
        ("Dopracowanie UX i czytelności", 16, ""),
        ("Konfiguracja TestFlight + build dystrybucyjny", 16, ""),
        ("Dokumentacja użytkownika / instruktora", 12, ""),
    ]),
]
mod_sum_rows = []
zebra = False
for mname, acts in v3:
    # naglowek modulu
    s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = s.cell(r, 1, mname); c.font = WHITE_BOLD; c.fill = fill(STEEL); c.alignment = LEFT
    s.row_dimensions[r].height = 18
    r += 1
    mfirst = r
    for act, godz, uw in acts:
        s.cell(r, 1, "").border = BORDER
        s.cell(r, 2, act).font = NORM; s.cell(r, 2).alignment = LEFT_T
        s.cell(r, 3, godz).alignment = CENTER; s.cell(r, 3).number_format = HRS
        s.cell(r, 4, f"=C{r}*Stawka").alignment = CENTER; s.cell(r, 4).number_format = PLN
        s.cell(r, 5, uw).font = SMALL; s.cell(r, 5).alignment = LEFT_T
        for j in range(1, 6):
            s.cell(r, j).border = BORDER
            if zebra: s.cell(r, j).fill = fill(GREY)
        r += 1
    mlast = r - 1
    # subtotal modulu
    s.cell(r, 2, f"Razem {mname.split(' — ')[0]}").font = BOLD
    s.cell(r, 2).alignment = LEFT
    s.cell(r, 3, f"=SUM(C{mfirst}:C{mlast})").font = BOLD; s.cell(r, 3).alignment = CENTER; s.cell(r, 3).number_format = HRS
    s.cell(r, 4, f"=SUM(D{mfirst}:D{mlast})").font = BOLD; s.cell(r, 4).alignment = CENTER; s.cell(r, 4).number_format = PLN
    for j in range(1, 6):
        s.cell(r, j).fill = fill(LIGHT); s.cell(r, j).border = BORDER
    mod_sum_rows.append(r)
    r += 1
    zebra = False

# RAZEM moduly
s.cell(r, 2, "RAZEM MODUŁY (development + testy)").font = BOLD; s.cell(r, 2).alignment = LEFT
gsum = "+".join([f"C{x}" for x in mod_sum_rows])
ksum = "+".join([f"D{x}" for x in mod_sum_rows])
s.cell(r, 3, f"={gsum}").font = BOLD; s.cell(r, 3).alignment = CENTER; s.cell(r, 3).number_format = HRS
s.cell(r, 4, f"={ksum}").font = BOLD; s.cell(r, 4).alignment = CENTER; s.cell(r, 4).number_format = PLN
for j in range(1, 6):
    s.cell(r, j).fill = fill(GREEN); s.cell(r, j).border = BORDER
modtot = r
r += 1
# Bufor
s.cell(r, 2, "Bufor ryzyka / rezerwa").font = NORM; s.cell(r, 2).alignment = LEFT
s.cell(r, 3, f"=ROUND(C{modtot}*Bufor_proc,0)").alignment = CENTER; s.cell(r, 3).number_format = HRS
s.cell(r, 4, f"=C{r}*Stawka").alignment = CENTER; s.cell(r, 4).number_format = PLN
s.cell(r, 5, "15% × godziny modułów").font = SMALL
for j in range(1, 6): s.cell(r, j).border = BORDER
buf = r; r += 1
# RAZEM PRACA
s.cell(r, 2, "RAZEM PRACA (Wariant A)").font = BOLD; s.cell(r, 2).alignment = LEFT
s.cell(r, 3, f"=C{modtot}+C{buf}").font = BOLD; s.cell(r, 3).alignment = CENTER; s.cell(r, 3).number_format = HRS
s.cell(r, 4, f"=D{modtot}+D{buf}").font = BOLD; s.cell(r, 4).alignment = CENTER; s.cell(r, 4).number_format = PLN
for j in range(1, 6): s.cell(r, j).fill = fill(GOLD); s.cell(r, j).border = BORDER
r += 2

# Warianty A/B/C zestawienie
r = section(s, "WARIANTY V3 — porównanie (z buforem ryzyka)", r, last_col=5)
r = col_header(s, r, ["Wariant", "Opis", "Godziny RAZEM", "Robocizna (PLN)", "Uwagi"])
v3_var = [
    ("A — Podstawowy (1:1)", "Własny kod, pełny zakres M1–M12", 1076, 0, "Bazowy wariant tego arkusza"),
    ("B — + Opcja 1:N", "Wariant A + połączenie 1 Pilot → do 4 Monitorów", 1226, 0, "Opcja 1:N wliczona"),
    ("C — + reuse II Library", "Wariant A z ponownym użyciem Infirmary Integrated", 948, 0, "Patrz korekty reuse niżej / arkusz V4"),
]
for war, opis, godz, _kwota, note in v3_var:
    s.cell(r, 1, war).font = BOLD; s.cell(r, 1).alignment = LEFT
    s.cell(r, 2, opis).font = NORM; s.cell(r, 2).alignment = LEFT_T
    s.cell(r, 3, godz).alignment = CENTER; s.cell(r, 3).number_format = HRS
    s.cell(r, 4, f"=C{r}*Stawka").alignment = CENTER; s.cell(r, 4).number_format = PLN; s.cell(r, 4).font = BOLD
    s.cell(r, 4).fill = fill(GOLD)
    s.cell(r, 5, note).font = SMALL; s.cell(r, 5).alignment = LEFT_T
    for j in range(1, 6): s.cell(r, j).border = BORDER
    r += 1
r += 1

# Korekty reuse (Wariant C)
r = section(s, "KOREKTY REUSE — Wariant C (Infirmary Integrated, Apache 2.0)", r, last_col=5)
r = col_header(s, r, ["Moduł", "Godz. własny (A)", "Korekta reuse", "Godz. Wariant C", "Uwaga"])
v3_reuse = [
    ("M1 — Architektura i P2P", 104, 0, "Własna warstwa — bez reuse"),
    ("M2 — Widok Monitora", 120, -40, "Reuse silnika fal i przebiegów"),
    ("M3 — Pilot: pulpit", 76, 0, "Własny UI"),
    ("M4 — Arytmie i zakłócenia", 130, -60, "Reuse 30+ rytmów (największa oszczędność)"),
    ("M5 — System scenariuszy", 102, -20, "Reuse modelu danych scenariusza"),
    ("M6 — Tryb Defib", 74, -10, "Reuse logiki defibrylacji"),
    ("M7 — Tryb AED", 54, -8, "Reuse logiki AED"),
    ("M8 — Kardiowersja", 44, -8, "Reuse logiki kardiowersji"),
    ("M9 — Stymulacja", 52, -10, "Reuse logiki pacingu"),
    ("M10 — Logika wyładowań", 52, 0, "Własny silnik reguł"),
    ("M11 — Ustawienia", 40, 0, "Własne"),
    ("M12 — Testy i dystrybucja", 88, 0, "Bez zmian"),
    ("Port II Library → MAUI", 0, 40, "Usunięcie zależności desktopowych, refaktor"),
    ("Obsługa licencji Apache 2.0", 0, 4, "Atrybucja, plik NOTICE, zgodność"),
]
first = r; zebra = False
for nm, ga, kor, uw in v3_reuse:
    s.cell(r, 1, nm).font = NORM; s.cell(r, 1).alignment = LEFT
    s.cell(r, 2, ga).alignment = CENTER; s.cell(r, 2).number_format = HRS
    s.cell(r, 3, kor).alignment = CENTER; s.cell(r, 3).number_format = '+0;-0;0'
    s.cell(r, 4, f"=B{r}+C{r}").alignment = CENTER; s.cell(r, 4).number_format = HRS; s.cell(r, 4).font = BOLD
    s.cell(r, 5, uw).font = SMALL; s.cell(r, 5).alignment = LEFT_T
    for j in range(1, 6):
        s.cell(r, j).border = BORDER
        if zebra: s.cell(r, j).fill = fill(GREY)
    zebra = not zebra
    r += 1
last = r - 1
s.cell(r, 1, "RAZEM moduły — Wariant C").font = BOLD; s.cell(r, 1).alignment = LEFT
s.cell(r, 2, f"=SUM(B{first}:B{last})").font = BOLD; s.cell(r, 2).alignment = CENTER; s.cell(r, 2).number_format = HRS
s.cell(r, 3, f"=SUM(C{first}:C{last})").font = BOLD; s.cell(r, 3).alignment = CENTER; s.cell(r, 3).number_format = '+0;-0;0'
s.cell(r, 4, f"=SUM(D{first}:D{last})").font = BOLD; s.cell(r, 4).alignment = CENTER; s.cell(r, 4).number_format = HRS
for j in range(1, 6): s.cell(r, j).fill = fill(GREEN); s.cell(r, j).border = BORDER
s.freeze_panes = "A5"

# ============================================================================
# V4 — Zunifikowany (REKOMENDOWANA)
# ============================================================================
s = setup_sheet("V4 — Zunifikowany ★", [7, 34, 13, 13, 14, 34])
title_bar(s, "WYCENA V4 — ZUNIFIKOWANA ★ (reuse + bufor 15% + koszty sprzętu)", last_col=6, fillc="385723")
s.merge_cells("A2:F2")
s.cell(2, 1, "REKOMENDOWANA, pełna oferta · zakres V3 + oszczędność reuse + bufor ryzyka + sprzęt · kolejność realizacji M1→M12").font = SMALL

r = 4
r = col_header(s, r, ["Nr", "Kamień milowy / moduł", "Godz. własny (A)", "Korekta reuse", "Godz. V4", "Koszt V4 (PLN) / Uwaga"])
v4_mods = [
    ("M1", "P2P i architektura", 104, 0, "Własna warstwa Multipeer; największe ryzyko techniczne"),
    ("M2", "Widok Monitora (przebiegi)", 120, -40, "Reuse silnika fal z II Library; integracja SkiaSharp/MAUI"),
    ("M3", "Pilot: pulpit instruktora", 76, 0, "Własny UI instruktora"),
    ("M4", "Arytmie i zakłócenia", 130, -60, "Reuse 30+ rytmów; weryfikacja wizualna i dopasowanie MAUI"),
    ("M5", "System scenariuszy", 102, -20, "Reuse modelu danych; własny edytor CRUD"),
    ("M6", "Tryb Defib", 74, -10, "Reuse logiki defibrylacji"),
    ("M7", "Tryb AED", 54, -8, "Reuse logiki AED"),
    ("M8", "Kardiowersja zsynchronizowana", 44, -8, "Reuse logiki kardiowersji"),
    ("M9", "Stymulacja (pacing)", 52, -10, "Reuse logiki pacingu; model uproszczony"),
    ("M10", "Logika wyładowań", 52, 0, "Własny silnik reguł"),
    ("M11", "Ustawienia i audio", 40, 0, "Własne ustawienia, audio, metronom"),
    ("M12", "Testy i dystrybucja", 88, 0, "Testy P2P na fizycznych iPadach; TestFlight"),
    ("+", "Port II Library → MAUI", 0, 40, "Usunięcie zależności desktopowych, refaktor"),
    ("+", "Obsługa licencji Apache 2.0", 0, 4, "Atrybucja, plik NOTICE, zgodność"),
]
first = r; zebra = False
for nr, nm, ga, kor, uw in v4_mods:
    s.cell(r, 1, nr).font = BOLD; s.cell(r, 1).alignment = CENTER
    s.cell(r, 2, nm).font = BOLD; s.cell(r, 2).alignment = LEFT_T
    s.cell(r, 3, ga).alignment = CENTER; s.cell(r, 3).number_format = HRS
    s.cell(r, 4, kor).alignment = CENTER; s.cell(r, 4).number_format = '+0;-0;0'
    s.cell(r, 5, f"=C{r}+D{r}").alignment = CENTER; s.cell(r, 5).number_format = HRS; s.cell(r, 5).font = BOLD
    s.cell(r, 6, uw).font = SMALL; s.cell(r, 6).alignment = LEFT_T
    for j in range(1, 7):
        s.cell(r, j).border = BORDER
        if zebra: s.cell(r, j).fill = fill(GREY)
    s.row_dimensions[r].height = 28
    zebra = not zebra
    r += 1
last = r - 1
# SUMA modulow
s.cell(r, 2, "SUMA MODUŁÓW").font = BOLD; s.cell(r, 2).alignment = LEFT
s.cell(r, 3, f"=SUM(C{first}:C{last})").font = BOLD; s.cell(r, 3).alignment = CENTER; s.cell(r, 3).number_format = HRS
s.cell(r, 5, f"=SUM(E{first}:E{last})").font = BOLD; s.cell(r, 5).alignment = CENTER; s.cell(r, 5).number_format = HRS
s.cell(r, 6, f"=E{r}*Stawka").font = BOLD; s.cell(r, 6).alignment = CENTER; s.cell(r, 6).number_format = PLN
for j in range(1, 7): s.cell(r, j).fill = fill(GREEN); s.cell(r, j).border = BORDER
modtot = r; r += 1
# Bufor
s.cell(r, 2, "Bufor ryzyka (15%)").font = NORM; s.cell(r, 2).alignment = LEFT
s.cell(r, 5, f"=ROUND(E{modtot}*Bufor_proc,0)").alignment = CENTER; s.cell(r, 5).number_format = HRS
s.cell(r, 6, f"=E{r}*Stawka").alignment = CENTER; s.cell(r, 6).number_format = PLN
for j in range(1, 7): s.cell(r, j).border = BORDER
buf = r; r += 1
# RAZEM godziny / robocizna
s.cell(r, 2, "RAZEM GODZINY / ROBOCIZNA").font = BOLD; s.cell(r, 2).alignment = LEFT
s.cell(r, 5, f"=E{modtot}+E{buf}").font = BOLD; s.cell(r, 5).alignment = CENTER; s.cell(r, 5).number_format = HRS
s.cell(r, 6, f"=E{r}*Stawka").font = BOLD; s.cell(r, 6).alignment = CENTER; s.cell(r, 6).number_format = PLN
for j in range(1, 7): s.cell(r, j).fill = fill(GOLD); s.cell(r, j).border = BORDER
robocizna = r; r += 2

# Koszty stale
r = section(s, "KOSZTY STAŁE (sprzęt, konta, licencje)", r, last_col=6)
r = col_header(s, r, ["", "Pozycja", "Ilość", "—", "Koszt (PLN)", "Uwaga"])
v4_fixed = [
    ("Mac mini M4 (host budowania iOS)", 0, 0, "Klient posiada — 0 PLN. Bez Maca: ~2 500–3 500 PLN."),
    ("Apple Developer Program (99 USD/rok)", 1, "=ROUND(99*KursUSD,0)", "Wymagany do TestFlight i podpisywania"),
    ("2 × iPad 10. gen (testy P2P)", 2, 5000, "Min. 2 do testów komunikacji P2P"),
    ("Akcesoria (etui, kable, ładowarki)", 1, 300, "Szacunkowo"),
]
first = r; zebra = False
for nm, il, koszt, uw in v4_fixed:
    s.cell(r, 2, nm).font = NORM; s.cell(r, 2).alignment = LEFT_T
    s.cell(r, 3, il).alignment = CENTER
    kc = s.cell(r, 5, koszt); kc.alignment = CENTER; kc.number_format = PLN
    s.cell(r, 6, uw).font = SMALL; s.cell(r, 6).alignment = LEFT_T
    for j in range(1, 7):
        s.cell(r, j).border = BORDER
        if zebra: s.cell(r, j).fill = fill(GREY)
    zebra = not zebra
    r += 1
last = r - 1
s.cell(r, 2, "RAZEM koszty stałe").font = BOLD; s.cell(r, 2).alignment = LEFT
s.cell(r, 5, f"=SUM(E{first}:E{last})").font = BOLD; s.cell(r, 5).alignment = CENTER; s.cell(r, 5).number_format = PLN
for j in range(1, 7): s.cell(r, j).fill = fill(LIGHT); s.cell(r, j).border = BORDER
fixed = r; r += 1

# SUMA CALKOWITA
s.cell(r, 2, "★ SUMA CAŁKOWITA (robocizna + sprzęt)").font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
s.cell(r, 2).alignment = LEFT
s.cell(r, 5, f"=F{robocizna}+E{fixed}").font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
s.cell(r, 5).alignment = CENTER; s.cell(r, 5).number_format = PLN
for j in range(1, 7): s.cell(r, j).fill = fill("385723"); s.cell(r, j).border = BORDER
s.row_dimensions[r].height = 24
s.freeze_panes = "A5"

# ============================================================================
# OPCJE DODATKOWE
# ============================================================================
s = setup_sheet("Opcje dodatkowe", [10, 50, 12, 14, 16, 30])
title_bar(s, "OPCJE DODATKOWE — praca poza zakresem podstawowym", last_col=6)
s.merge_cells("A2:F2")
s.cell(2, 1, "NIE wliczane do sum głównych. Koszt 'z buforem' = wersja doliczona do wariantów V3/V4 (bufor ryzyka 15%).").font = SMALL
r = 4
r = col_header(s, r, ["Opcja", "Zakres", "Godziny", "Koszt bazowy (PLN)", "Z buforem 15% (PLN)", "Uwagi"])
opts = [
    ("A", "Połączenie 1:N (1 Pilot → do 4 Monitorów): broadcast, sesje per-urządzenie, pulpit master, testy 4+ urządzeń, bufor multi-peer", 130, True, "Wymaga dodatkowych iPadów (testy)"),
    ("B", "Pełny model capture/no-capture w Pacingu (próg wychwycenia mA/rate) zamiast modelu uproszczonego", 24, True, "Alternatywa dla uproszczonego M9"),
    ("C", "Port na Android: natywna warstwa P2P (Nearby Connections), testy na sprzęcie, zgodność layoutów", 150, False, "Orientacyjnie 120–180 h (~7 500–9 000 PLN)"),
]
zebra = False
for op, zk, godz, withbuf, uw in opts:
    s.cell(r, 1, "Opcja " + op).font = BOLD; s.cell(r, 1).alignment = CENTER
    s.cell(r, 2, zk).font = NORM; s.cell(r, 2).alignment = LEFT_T
    s.cell(r, 3, godz).alignment = CENTER; s.cell(r, 3).number_format = HRS
    s.cell(r, 4, f"=C{r}*Stawka").alignment = CENTER; s.cell(r, 4).number_format = PLN
    if withbuf:
        s.cell(r, 5, f"=ROUND(D{r}*(1+Bufor_proc),0)").alignment = CENTER; s.cell(r, 5).number_format = PLN; s.cell(r, 5).fill = fill(GOLD)
    else:
        s.cell(r, 5, "~7 500–9 000").alignment = CENTER
    s.cell(r, 6, uw).font = SMALL; s.cell(r, 6).alignment = LEFT_T
    for j in range(1, 7):
        s.cell(r, j).border = BORDER
        if zebra and j != 5: s.cell(r, j).fill = fill(GREY)
    s.row_dimensions[r].height = 42
    zebra = not zebra
    r += 1
s.freeze_panes = "A5"

# ============================================================================
# KOSZTY STAŁE (pelny przeglad sprzetu i dystrybucji)
# ============================================================================
s = setup_sheet("Koszty stałe", [38, 18, 14, 50])
title_bar(s, "KOSZTY STAŁE — sprzęt, konta, dystrybucja", last_col=4)
s.merge_cells("A2:D2")
s.cell(2, 1, "Identyczne dla wszystkich wariantów. Mac założono jako posiadany przez klienta.").font = SMALL
r = 4
r = section(s, "MAC — wymagany do budowania aplikacji iOS", r, last_col=4)
r = col_header(s, r, ["Model", "Cena (PLN)", "—", "Uwagi"])
macs = [
    ("Mac mini M4", 2500, "Najtańsza opcja, wymaga monitora/klawiatury"),
    ("MacBook Air M3 13\"", 5500, "Najpopularniejszy wybór dewelopera"),
    ("MacBook Pro M3 14\"", 8500, "Nadmiarowy do tego projektu"),
]
for m, c, uw in macs:
    s.cell(r, 1, m).font = NORM; s.cell(r, 1).alignment = LEFT
    s.cell(r, 2, c).alignment = CENTER; s.cell(r, 2).number_format = PLN
    s.cell(r, 4, uw).font = SMALL; s.cell(r, 4).alignment = LEFT_T
    for j in range(1, 5): s.cell(r, j).border = BORDER
    r += 1
r += 1
r = section(s, "iPADY DO TESTÓW (wymagane tylko do testów P2P)", r, last_col=4)
r = col_header(s, r, ["Model iPada", "Cena (PLN/szt.)", "—", "Uwagi"])
ipads = [
    ("iPad 10. generacji", 2200, "Minimum do testów"),
    ("iPad Air M2", 3200, "Komfortowe urządzenie testowe"),
    ("2× iPad (testy P2P)", 4400, "Minimalne — 2× iPad 10. gen"),
]
for m, c, uw in ipads:
    s.cell(r, 1, m).font = NORM; s.cell(r, 1).alignment = LEFT
    s.cell(r, 2, c).alignment = CENTER; s.cell(r, 2).number_format = PLN
    s.cell(r, 4, uw).font = SMALL; s.cell(r, 4).alignment = LEFT_T
    for j in range(1, 5): s.cell(r, j).border = BORDER
    r += 1
r += 1
r = section(s, "APPLE DEVELOPER ACCOUNT / DYSTRYBUCJA", r, last_col=4)
r = col_header(s, r, ["Model dystrybucji", "Koszt roczny (PLN)", "—", "Uwagi"])
dist = [
    ("App Store (publiczny) — Apple Developer Program", 400, "99 USD/rok"),
    ("TestFlight — Apple Developer Program", 400, "Rekomendowane dla aplikacji okołomedycznej"),
    ("Enterprise — Apple Developer Enterprise Program", 1200, "299 USD/rok, dystrybucja wewnętrzna"),
]
for m, c, uw in dist:
    s.cell(r, 1, m).font = NORM; s.cell(r, 1).alignment = LEFT
    s.cell(r, 2, c).alignment = CENTER; s.cell(r, 2).number_format = PLN
    s.cell(r, 4, uw).font = SMALL; s.cell(r, 4).alignment = LEFT_T
    for j in range(1, 5): s.cell(r, j).border = BORDER
    r += 1
s.freeze_panes = "A5"

# ============================================================================
# RYZYKA
# ============================================================================
s = setup_sheet("Ryzyka", [6, 46, 14, 14, 52])
title_bar(s, "REJESTR RYZYK I ZAGROŻEŃ", last_col=5)
r = 3
r = col_header(s, r, ["ID", "Ryzyko", "Prawdopodobieństwo", "Wpływ", "Działanie ograniczające (mitygacja)"])
risks = [
    ("R1", "Multipeer Connectivity w .NET MAUI wymaga własnego bindingu warstwy natywnej", "Wysokie", "Wysoki", "PoC w M1 jako pierwszy krok; plan B: TCP/UDP lokalnie. Bufor 15%."),
    ("R2", "Brak doświadczenia z Apple/iOS (Xcode, podpisywanie, provisioning)", "Wysokie", "Średni", "Materiały pomocnicze; TestFlight zamiast App Store ogranicza tarcia."),
    ("R3", "Wiarygodność wizualna przebiegów EKG/arytmii trudniejsza niż zakładano", "Średnie", "Wysoki", "Najpierw silnik (M2), potem stany (M4); reuse (Wariant C/V4) mocno ogranicza."),
    ("R4", "Uwagi/odrzucenie Apple (aplikacja „okołomedyczna”)", "Średnie", "Średni", "TestFlight/wewnętrznie; oznaczenie „symulacja edukacyjna, nie wyrób medyczny”."),
    ("R5", "Stabilność synchronizacji w czasie rzeczywistym (opóźnienia, desync)", "Średnie", "Wysoki", "Prosty, odporny protokół; reconnect; testy na realnym sprzęcie."),
    ("R6", "Skalowanie 1:N (do 4) — wydajność i stabilność Multipeer", "Średnie", "Średni", "Bufor w Opcji A; testy z 4+ urządzeniami; limit 4 klientów."),
    ("R7", "Brak fizycznych iPadów = brak realnych testów P2P", "Wysokie", "Wysoki", "Zakup min. 2 iPadów (Koszty stałe) na starcie."),
    ("R8", "Niedoszacowanie zakresu („pełna wersja, nie MVP”)", "Średnie", "Wysoki", "Kamienie milowe; przyrostowe dostarczanie; rezerwa 15%."),
    ("R9", "Zależność od jednego programisty", "Średnie", "Wysoki", "Dokumentacja na bieżąco; repo pod kontrolą wersji."),
    ("R10", "Złożoność audio (AED, metronom, alarmy) na iPadOS", "Niskie", "Niski", "Standardowe API audio; wczesny prototyp."),
    ("R11", "[Wariant C/V4] Zależność od kodu II Library; zgodność z Apache 2.0", "Średnie", "Średni", "Czas na analizę i port; dołączenie licencji/NOTICE; możliwość fork-a."),
]
zebra = False
def lvl_color(v):
    return {"Wysokie": "F8CBAD", "Wysoki": "F8CBAD", "Średnie": "FFE699",
            "Średni": "FFE699", "Niskie": "C6E0B4", "Niski": "C6E0B4"}.get(v, "FFFFFF")
for rid, ryz, pr, wp, mit in risks:
    s.cell(r, 1, rid).font = BOLD; s.cell(r, 1).alignment = CENTER
    s.cell(r, 2, ryz).font = NORM; s.cell(r, 2).alignment = LEFT_T
    s.cell(r, 3, pr).alignment = CENTER; s.cell(r, 3).fill = fill(lvl_color(pr))
    s.cell(r, 4, wp).alignment = CENTER; s.cell(r, 4).fill = fill(lvl_color(wp))
    s.cell(r, 5, mit).font = NORM; s.cell(r, 5).alignment = LEFT_T
    for j in range(1, 6): s.cell(r, j).border = BORDER
    s.row_dimensions[r].height = 32
    r += 1
s.freeze_panes = "A4"

# ----------------------------------------------------------------------------
out = "Wycena_Kielich_ZBIORCZA_v2.xlsx"
wb.save(out)
print("Zapisano:", out)
print("Arkusze:", wb.sheetnames)
